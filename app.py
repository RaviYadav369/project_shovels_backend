from flask import Flask, request, jsonify
from flask_cors import CORS, cross_origin
import os
from pymongo import MongoClient
from svix.api import Svix 
from dotenv import load_dotenv
from modules.mongodb import (
    save_email_db, login_user, save_document, get_document, delete_document, 
    register_user, get_user, get_all_users, save_transcripts, save_channel, 
    get_channels, delete_channel_db, add_session, get_all_sessions, get_session, 
    delete_session, get_history, update_user, delete_user,register_user    
)
from modules.ai_tools import get_video_links, make_transcript
from werkzeug.utils import secure_filename
import uuid
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from modules.chatbot import get_chatbot_response_agent
import json
from svix.webhooks import Webhook, WebhookVerificationError
import os
import json
import hmac
import base64
import hashlib
from datetime import datetime

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY')
WEBHOOK_SECRET = os.getenv("WEBHOOK_SECRET").encode()

# MongoDB Setup
MONGO_URI = os.getenv("MONGO_URI")
client = MongoClient(MONGO_URI)
db = client['yt_chatbot']
users_collection = db['users']

CORS(app, resources={r"/*": {"origins": "*", "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"], "allow_headers": ["Authorization", "Content-Type"]}})

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('user', {}).get('role') != 'admin':
            return jsonify({'success': False, 'message': 'Access denied'})
        return f(*args, **kwargs)
    return decorated_function

@app.route('/', methods=['GET'])
def index():
    return jsonify({"message": "Welcome to the Shovel YT Chatbot API"})

@app.route("/webhook", methods=["POST"])
# def handle_webhook():
#     try:
#         headers = request.headers
#         WEBHOOK_SECRET = os.environ["WEBHOOK_SECRET"]

#         if not WEBHOOK_SECRET:
#             raise Exception("Please add WEBHOOK_SECRET from Clerk Dashboard to .env or .env.local")

#         # svix_id = headers.get("Svix-Id")
#         # svix_timestamp = headers.get("Svix-Timestamp")
#         # svix_signature = headers.get("Svix-Signature")
#         svix_id = headers.get("svix-id")
#         svix_timestamp = headers.get("svix-timestamp")
#         svix_signature = headers.get("svix-signature")
#         user_agent = headers.get("User-Agent")

#         if not svix_id or not svix_timestamp or not svix_signature:
#             return jsonify({'message':"Error occured -- no svix headers", 'status_code':400})

#         print('svix ids',svix_id, svix_timestamp, svix_signature,user_agent)
#         body = request.get_json()
#         body = json.dumps(body)
#         print("BoDy",body)
#         print("Headers",headers)
        
#         try:
#             wh = Webhook(WEBHOOK_SECRET)
#             evt = wh.verify(data=body,headers=headers)
#         except Exception as e:
#             print(f"Error verifying webhook: {e}")
#             return jsonify({'message':"Error occured in verifying", 'status_code':400})
    
#         event_type = evt.type
#         print("veriFicaTion")
#         print(evt)
        
#         if event_type == "user.created":
#             # Extract relevant information
#             clerk_id = evt.data["id"]
#             email = evt.data["email_addresses"][0]["email_address"]
#             first_name = evt.data["first_name"]
#             last_name = evt.data["last_name"]
#             profile_image_url = evt.data.get("profile_image_url", "")

#             # Use the register_user function to insert/update the user in the database
#             user_id = register_user(clerk_id, email, first_name, last_name, profile_image_url)
            
#             return jsonify({'message': "User created event processed successfully", 'user_id': user_id, 'status_code': 200})
#     except Exception as e:
#           print("Error Occured")
#           return jsonify({'message': "Error occurred", 'status_code' : 400})
       
# class Webhook:
#     def _init_(self, secret):
#         self.secret = secret.encode("utf-8")

#     def verify(self, body, headers):
#         data = json.dumps(body, sort_keys=True).encode("utf-8")

#         timestamp = int(headers["svix-timestamp"])

#         if (datetime.now().timestamp() - timestamp) > 60:
#             raise Exception("Timestamp is too old")

#         message = f"{data}\n{timestamp}\n{self.secret}"

#         signature = hmac.new(
#             self.secret,
#             message.encode("utf-8"),
#             hashlib.sha256
#         ).hexdigest()

#         if signature != headers["svix-signature"]:
#             raise Exception("Invalid signature")

#         return json.loads(body)

def handle_webhook():
    try:
        headers = request.headers
        WEBHOOK_SECRET = os.environ["WEBHOOK_SECRET"]

        if not WEBHOOK_SECRET:
            raise Exception("Please add WEBHOOK_SECRET from Clerk Dashboard to .env or .env.local")

        svix_id = headers.get("svix-id")
        svix_timestamp = headers.get("svix-timestamp")
        svix_signature = headers.get("svix-signature")
        user_agent = headers.get("User-Agent")

        if not svix_id or not svix_timestamp or not svix_signature:
            return jsonify({'message':"Error occured -- no svix headers", 'status_code':400})

        print('svix ids',svix_id, svix_timestamp, svix_signature,user_agent)
        body = request.get_json()
        body = json.dumps(body)
        print("BoDy",body)
        print("Headers",headers)
        
        try:
            wh = Webhook(WEBHOOK_SECRET)
            evt = wh.verify(data=body,headers=headers)
        except WebhookVerificationError as e:
            print(f"Error verifying webhook: {e}")
            return jsonify({'message':"Error occured in verifying", 'status_code':400})
    
        event_type = evt.type
        print("veriFicaTion",event_type)
        print(evt)
        
        # if event_type == "user.created":
        #     # Extract relevant information
        #     clerk_id = evt.data["id"]
        #     email = evt.data["email_addresses"][0]["email_address"]
        #     first_name = evt.data["first_name"]
        #     last_name = evt.data["last_name"]
        #     profile_image_url = evt.data.get("profile_image_url", "")

        #     # Use the register_user function to insert/update the user in the database
        #     user_id = register_user(clerk_id, email, first_name, last_name, profile_image_url)
            
        #     return jsonify({'message': "User created event processed successfully", 'user_id': user_id, 'status_code': 200})
    except Exception as e:
          print("Error Occured",e)
          return jsonify({'message': "Error occurred", 'status_code' : 400})
       
class Webhook:
    def __init__(self, secret):
        self.secret = secret.encode("utf-8")

    def verify(self, data, headers):
        svix_id = headers.get("svix-id")
        svix_timestamp = headers.get("svix-timestamp")

        signed_content = f"{svix_id}.{svix_timestamp}.{data}"

        # Decode the secret from base64
        secret_bytes = base64.b64decode(self.secret.split("_")[1])

        # Compute the expected signature
        signature = hmac.new(
            secret_bytes,
            signed_content.encode("utf-8"),
            hashlib.sha256
        ).digest()

        # Convert the expected signature to base64
        expected_signature = base64.b64encode(signature).decode("utf-8")
        print("Expected Signature",expected_signature,headers["svix-signature"])

        # Compare the expected signature to the actual signature
        if expected_signature!= headers["svix-signature"]:
            raise WebhookVerificationError("Invalid signature")

        return json.loads(data)

# login
@app.route('/login', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def login():
    # take email and password from request
    email = request.json['email']
    password = request.json['password']
    # check if user exists
    user = login_user(email, password)
    if user:
        # set session
        session['user'] = user
        return jsonify({'success': True, 'user': user[0], 'role': user[1]})
    else:
        return jsonify({'success': False})
    
# logout
@app.route('/logout', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def logout():
    session.pop('user', None)
    return jsonify({'success': True})

@app.route('/register', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def register():
    # take email and password from request
    email = request.json['email']
    password = request.json['password']
    # register user
    user = register_user(email, password)

    if user:
        # set session
        session['user'] = user
        return jsonify({'success': True, 'user': user[0],'role':user[1]})
    else:
        return jsonify({'success': False, 'message': 'User already exists'})
    

# get user details based on user id in body
@app.route('/user', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def user():
    # take email and password from request
    user_id = request.json['user_id']
    # check if user exists
    user = get_user(user_id)
    if user:
        return jsonify({'success': True, 'user': user})
    else:
        return jsonify({'success': False})
    

    # admin can get all users
@app.route('/admin/users', methods=['GET'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def users():
    # get all
    users = get_all_users()
    return jsonify({'success': True, 'users': users})
    
# add youtube channel
@app.route('/add_channel', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def add_channel():
    # take channel url from request
    channel_url = request.json['channel_url']
    user_id = request.json['user_id']

    # save channelurl in db
    channel_id=save_channel(user_id,channel_url)

    # get video links
    try:
        links = get_video_links(str(channel_url))
        links=links[:5]
        # make transcript
        print("generating transcript for channel")
        for i in links:
            transcript = make_transcript(i)
            ans=save_transcripts(user_id,transcript,str(channel_id))
        print("transcript generated")
        add_chat_in_db=add_session(user_id,channel_id)
        if add_chat_in_db:
            print("session added")
    except Exception as e:
        print(f"Error adding channel: {e}")
        return jsonify({'success': False})
    # return channel id
    return jsonify({'success': True, 'channel_id': str(channel_id)})
    
    # list of channels for a user
@app.route('/channels', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def channels():
    # take user id from request
    user_id = request.json['user_id']
    # get channels
    channels = get_channels(user_id)
    return jsonify({'success': True, 'channels': channels})


# delete channel and all itss transcripts
@app.route('/delete_channel', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def delete_channel():
    # take channel id from request
    channel_id = request.json['channel_id']
    # delete channel
    success = delete_channel_db(channel_id)
    return jsonify({'success': success})

#  get all chats for a user
@app.route('/sessions', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def chats():
    # take user id from request
    user_id = request.json['user_id']
    # get chats
    chats = get_all_sessions(user_id)
    chats = list(chats)
    for chat in chats:
        chat['_id'] = str(chat['_id'])
        chat['user_id'] = str(chat['user_id'])
        chat['channel_id'] = str(chat['channel_id'])
    return jsonify({'success': True, 'sessions': chats})

#  get chat for a user and channel
@app.route('/session', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def chat():
    # take user id from request
    user_id = request.json['user_id']
    channel_id = request.json['channel_id']
    # get chats
    chat = get_session(user_id,channel_id)
    return jsonify({'success': True, 'session': chat})

# new chat for a user without channel
@app.route('/start_new_session', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def new_chat():
    # take user id from request
    user_id = request.json['user_id']
    # random channel id*********************
    channel_id = str(uuid.uuid4())  # Generate a unique channel ID
    session_id=add_session(user_id,channel_id)
    return jsonify({'success': True, 'sessionId': session_id,'channelId':channel_id})

# delete chat for a user and channel
@app.route('/delete_session', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def delete_session_chat():
    # take user id from request
    session_id = request.json['session_id']
   
    # get chats
    success = delete_session(session_id)
    return jsonify({'success': success})


# chatbot response
@app.route('/chatbot', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def chatbot():
    # take user id from request
    data = request.json
    query = data.get('query')
    user_id = data.get('user_id')
    session_id = data.get('session_id')  # Use .get() to avoid KeyError
    channel_id = data.get('channel_id')
    if not session_id:
        return jsonify({'success': False, 'message': 'Session ID is missing'})


    # get chats
    chatbot_response = get_chatbot_response_agent(query, user_id,channel_id,session_id)
    return jsonify({'success': True ,'role':'assistant','response': chatbot_response})


#  get chat history based on user id and session id
@app.route('/history', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def history():
    # take user id from request
    user_id = request.json['user_id']
    session_id = request.json['session_id']
    # get chats
    history = get_history(session_id,user_id)
    return jsonify({'success': True, 'history': history})

#  get all users
@app.route('/users', methods=['GET'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def all_users():

    # get chats
    users = get_all_users()
    return jsonify({'success': True, 'users': users})

#  update user
@app.route('/update_user', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def update_user_api():
    # take user id from request
    user_id = request.json['user_id']
    email = request.json['email']
    password = request.json['password']
    role = request.json['role']
    # get chats
    success = update_user(user_id,email,password,role)
    return jsonify({'success': success})

#  delete user
@app.route('/delete_user', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def delete_user_api():
    # take user id from request
    user_id = request.json['user_id']
    # get chats
    success = delete_user(user_id)
    return jsonify({'success': success})
# save emails from comming soon page into the database
@app.route('/save_email', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def save_email():
    # take email from request
    email = request.json['email']
    # save email in db
    success = save_email_db(email)

    return jsonify({'success': success})

def extract_text_from_pdf(file):
    reader = PyPDF2.PdfReader(file)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text

def extract_text_from_docx(file):
    doc = Document(file)
    text = ""
    for para in doc.paragraphs:
        text += para.text + "\n"
    return text

def extract_text_from_xlsx(file):
    workbook = load_workbook(filename=file)
    sheet = workbook.active
    text = ""
    for row in sheet.iter_rows(values_only=True):
        text += " ".join(map(str, row)) + "\n"
    return text

def extract_text_from_file(file):
    # Determine the file type (e.g., from the file extension)
    file_type = file.filename.split('.')[-1].lower()
    if file_type == 'pdf':
        return extract_text_from_pdf(file)
    elif file_type == 'docx':
        return extract_text_from_docx(file)
    elif file_type == 'xlsx':
        return extract_text_from_xlsx(file)
    elif file_type == 'txt':
        return file.read().decode('utf-8')
    else:
        raise ValueError("Unsupported file type")

@app.route('/upload_document', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def upload_document():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    user_id = request.form.get('user_id')
    if not user_id:
        return jsonify({"error": "User ID is missing"}), 400
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    channel_id = str(uuid.uuid4())  # Generate a random channel ID

    # Process the file to extract text
    # This part depends on the file type (PDF, DOCX, etc.)
    text_content = extract_text_from_file(file)  # Implement this function based on file type
    filename=secure_filename(file.filename)
    # Save the extracted text as a transcript
    success = save_transcripts(user_id, text_content, channel_id)
    if not success:
        return jsonify({"error": "Failed to save transcript"}), 500
    trans=save_document(user_id, success,filename)
    add_chat_in_db=add_session(user_id,channel_id)
    if add_chat_in_db:
        print("session added")

    if success:
        return jsonify({"success": True, "channel_id": str(channel_id), "transcript_id": str(success)}), 200
    else:
        return jsonify({"error": "Failed to save transcript"}), 500
    
@app.route('/delete_document', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def delete_document_api():
    # take user id from request
    user_id = request.json['user_id']
    
    transcript_id = request.json['transcript_id']
    # get chats
    success = delete_document(user_id,transcript_id)
    if not success:
        return jsonify({"error": "Failed to delete document"}), 500
    return jsonify({'success': success})
    
# list_documents
@app.route('/documents', methods=['POST'])
@cross_origin(origin='*', headers=['Content-Type', 'Authorization'])
def documents():
    # take user id from request
    user_id = request.json['user_id']
    # get chats
    documents = get_document(user_id)

    return jsonify({'success': True, 'documents': documents})



if __name__ == '__main__':
    app.run(host='0.0.0.0',debug=True,threaded=True,use_reloader=False)
